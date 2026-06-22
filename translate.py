#!/usr/bin/env python3
"""
다국어 상세페이지 번역 도구
Naver CLOVA OCR + DeepL API

배치 모드 (권장):
    python translate.py
    → input/<제품명>/ 폴더를 모두 처리
    → output/<제품명>/<제품명>_translation.xlsx 생성

단일 이미지 모드:
    python translate.py input/제품명/이미지.jpg

번역 언어 설정: config.py의 TARGET_LANGS 수정
  예) TARGET_LANGS = ['JA', 'ZH', 'EN-US']
"""

import sys
import io
import os
import re
import tempfile
import shutil
import requests
import base64
import uuid
import time
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

import config

MAX_CHUNK_HEIGHT = 7500  # CLOVA OCR 최대 허용 높이 8000px
CROP_MAX_H = 70
CROP_MAX_W = 220
ROW_HEIGHT_PT = 55

# DeepL 언어 코드 → 표시명
LANG_NAMES = {
    'JA':      '일본어',
    'ZH':      '중국어(간체)',
    'ZH-HANT': '중국어(번체)',
    'EN-US':   '영어',
    'EN-GB':   '영어(영국)',
    'DE':      '독일어',
    'FR':      '프랑스어',
    'ES':      '스페인어',
    'IT':      '이탈리아어',
    'PT-BR':   '포르투갈어',
}

COLORS = [
    ('#E53E3E', 'E53E3E'),
    ('#DD6B20', 'DD6B20'),
    ('#D69E2E', 'D69E2E'),
    ('#38A169', '38A169'),
    ('#3182CE', '3182CE'),
    ('#805AD5', '805AD5'),
    ('#D53F8C', 'D53F8C'),
]

IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.tiff', '.tif'}


# ──────────────────────────────────────────
# 1. OCR
# ──────────────────────────────────────────

def _call_ocr_api(image_bytes: bytes, ext: str) -> list:
    """CLOVA OCR API 단일 호출 → fields 반환"""
    payload = {
        "images": [{"format": ext, "name": "image", "data": base64.b64encode(image_bytes).decode('utf-8')}],
        "requestId": str(uuid.uuid4()),
        "version": "V2",
        "timestamp": int(time.time() * 1000),
    }
    api_url = config.CLOVA_API_URL.replace('http://', 'https://')
    response = requests.post(
        api_url,
        headers={"X-OCR-SECRET": config.CLOVA_SECRET_KEY, "Content-Type": "application/json"},
        json=payload,
        timeout=60,
    )
    response.raise_for_status()
    fields = []
    for img_result in response.json().get('images', []):
        fields.extend(img_result.get('fields', []))
    return fields


def find_safe_cut(img: Image.Image, target_y: int, search_range: int = 300) -> int:
    """target_y 주변에서 픽셀 분산이 가장 낮은 행(빈 공간)을 찾아 반환"""
    height = img.height
    y_min = max(10, target_y - search_range)
    y_max = min(height - 10, target_y + search_range)

    # 탐색 영역만 잘라 한 번에 픽셀 로드 (L모드: 1바이트 = 1픽셀)
    strip = img.convert('L').crop((0, y_min, img.width, y_max))
    pixels = list(strip.tobytes())
    width = img.width
    strip_height = y_max - y_min

    best_y = target_y
    min_variance = float('inf')

    for i in range(strip_height):
        row = pixels[i * width:(i + 1) * width]
        mean = sum(row) / width
        variance = sum((p - mean) ** 2 for p in row) / width
        if variance < min_variance:
            min_variance = variance
            best_y = y_min + i

    return best_y


def ocr_image(img: Image.Image) -> list:
    """PIL 이미지를 청크로 분할하여 OCR, y좌표 보정 후 fields 반환
    슬라이스 경계는 텍스트가 없는 빈 공간에서 자름"""
    img = img.convert('RGB')
    width, height = img.size

    if height <= MAX_CHUNK_HEIGHT:
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=95)
        return _call_ocr_api(buf.getvalue(), 'jpeg')

    all_fields = []
    y_start, chunk_idx = 0, 0
    while y_start < height:
        target_end = y_start + MAX_CHUNK_HEIGHT

        if target_end >= height:
            y_end = height
        else:
            y_end = find_safe_cut(img, target_end, search_range=300)
            print(f"        슬라이스 기준점: {target_end}px → 안전 지점: {y_end}px")

        chunk = img.crop((0, y_start, width, y_end))
        buf = io.BytesIO()
        chunk.save(buf, format='JPEG', quality=95)
        print(f"        청크 {chunk_idx + 1} OCR... (y: {y_start}~{y_end}px)")
        fields = _call_ocr_api(buf.getvalue(), 'jpeg')
        for field in fields:
            for v in field['boundingPoly']['vertices']:
                v['y'] += y_start
        all_fields.extend(fields)
        y_start = y_end
        chunk_idx += 1

    return all_fields


def group_into_lines(fields: list) -> list:
    """fields를 줄 단위로 그룹화"""
    lines, current = [], []
    for idx, field in enumerate(fields):
        current.append(field)
        if field.get('lineBreak', False) or idx == len(fields) - 1:
            text = ' '.join(w['inferText'] for w in current).strip()
            if text:
                verts = []
                for w in current:
                    verts.extend(w['boundingPoly']['vertices'])
                xs = [v['x'] for v in verts]
                ys = [v['y'] for v in verts]
                lines.append({'text': text, 'x1': min(xs), 'y1': min(ys), 'x2': max(xs), 'y2': max(ys)})
            current = []
    return lines


_KOREAN_RE = re.compile(r'[가-힣ᄀ-ᇿ㄰-㆏]')
_SENTENCE_END = re.compile(r'[.。!！?？]\s*$')

def filter_korean_only(items: list) -> list:
    """한국어가 포함된 항목만 남김 (숫자·기호·영문만인 항목 제거)"""
    return [item for item in items if _KOREAN_RE.search(item['text'])]


def _merge_group(group: list) -> dict:
    if len(group) == 1:
        return group[0]
    return {
        'text': ' '.join(l['text'] for l in group),
        'x1': min(l['x1'] for l in group),
        'y1': group[0]['y1'],
        'x2': max(l['x2'] for l in group),
        'y2': group[-1]['y2'],
    }


def merge_paragraph_lines(lines: list) -> list:
    """문장 종결 부호 기반으로 연속 줄을 하나의 문장으로 병합.
    이전 줄이 .。!？로 끝나지 않으면 다음 줄과 합침.
    수직 간격이 평균 줄높이의 2배 초과 시 강제로 새 문장 시작."""
    if not lines:
        return []

    lines = sorted(lines, key=lambda l: l['y1'])

    avg_line_h = sum(l['y2'] - l['y1'] for l in lines) / len(lines)
    gap_threshold = avg_line_h * 2.0

    merged = []
    group = [lines[0]]

    for line in lines[1:]:
        prev = group[-1]
        gap = line['y1'] - prev['y2']
        sentence_ended = bool(_SENTENCE_END.search(prev['text']))

        if sentence_ended or gap > gap_threshold:
            merged.append(_merge_group(group))
            group = [line]
        else:
            group.append(line)

    merged.append(_merge_group(group))
    return merged


# ──────────────────────────────────────────
# 2. 번역
# ──────────────────────────────────────────

def translate_batch(texts: list, target_lang: str) -> list:
    """DeepL API로 한→target_lang 일괄 번역 (50개씩 청크)"""
    results = []
    chunk_size = 50
    for i in range(0, len(texts), chunk_size):
        chunk = texts[i:i + chunk_size]
        response = requests.post(
            "https://api-free.deepl.com/v2/translate",
            headers={"Authorization": f"DeepL-Auth-Key {config.DEEPL_API_KEY}"},
            json={'text': chunk, 'source_lang': 'KO', 'target_lang': target_lang},
            timeout=30,
        )
        response.raise_for_status()
        results.extend(t['text'] for t in response.json()['translations'])
    return results


def translate_all_langs(texts: list, target_langs=None) -> dict:
    """선택한 모든 언어로 번역 → {lang_code: [번역결과, ...]}
    target_langs 미지정 시 config.TARGET_LANGS 사용(CLI 폴백)."""
    if target_langs is None:
        target_langs = getattr(config, 'TARGET_LANGS', ['JA'])
    results = {}
    for lang in target_langs:
        lang_name = LANG_NAMES.get(lang, lang)
        print(f"    [{lang_name}] 번역 중...")
        results[lang] = translate_batch(texts, lang)
    return results


# ──────────────────────────────────────────
# 3. 이미지 처리
# ──────────────────────────────────────────

def combine_vertically(images: list) -> Image.Image:
    """여러 PIL 이미지를 세로로 합치기"""
    if len(images) == 1:
        return images[0].copy().convert('RGB')
    total_w = max(img.width for img in images)
    total_h = sum(img.height for img in images)
    combined = Image.new('RGB', (total_w, total_h), (255, 255, 255))
    y = 0
    for img in images:
        combined.paste(img.convert('RGB'), (0, y))
        y += img.height
    return combined


def create_annotated_image(pil_img: Image.Image, items: list) -> Image.Image:
    """PIL 이미지에 번호 박스만 오버레이 후 반환"""
    img = pil_img.copy().convert('RGB')
    draw = ImageDraw.Draw(img)
    badge_r = max(18, img.width // 60)
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", badge_r)
    except Exception:
        font = ImageFont.load_default()

    for i, item in enumerate(items):
        hex_color, _ = COLORS[i % len(COLORS)]
        x1, y1, x2, y2 = item['x1'], item['y1'], item['x2'], item['y2']
        draw.rectangle([x1, y1, x2, y2], outline=hex_color, width=3)
        bx, by = x1 + badge_r, y1
        draw.ellipse([bx - badge_r, by - badge_r, bx + badge_r, by + badge_r], fill=hex_color)
        draw.text((bx, by), str(i + 1), fill='white', font=font, anchor='mm')
    return img


def _make_crop(orig_img: Image.Image, item: dict, out_path: str):
    pad = 15
    x1 = max(0, item['x1'] - pad)
    y1 = max(0, item['y1'] - pad)
    x2 = min(orig_img.width, item['x2'] + pad)
    y2 = min(orig_img.height, item['y2'] + pad)
    crop = orig_img.crop((x1, y1, x2, y2))
    scale = min(CROP_MAX_W / max(crop.width, 1), CROP_MAX_H / max(crop.height, 1), 1.0)
    new_w, new_h = max(1, int(crop.width * scale)), max(1, int(crop.height * scale))
    crop = crop.resize((new_w, new_h), Image.LANCZOS)
    crop.save(out_path, quality=92)
    return new_w, new_h


# ──────────────────────────────────────────
# 4. Excel 생성
# ──────────────────────────────────────────

def create_excel(items: list, orig_img: Image.Image, annotated_img: Image.Image, output, target_langs=None):
    """전체 주석 이미지(좌 A-F열) + 번역표(우) — y좌표 기반 행 정렬.
    target_langs 의 언어 수만큼 번역 컬럼이 생성된다.
    output: 파일 경로(str) 또는 file-like 객체(BytesIO 등) 모두 허용."""
    if target_langs is None:
        target_langs = getattr(config, 'TARGET_LANGS', ['JA'])
    IMAGE_COLS = 6          # A-F열을 이미지 영역으로 사용
    PX_PER_ROW = ROW_HEIGHT_PT / 0.75  # 55pt → ~73.3px (pt→px 변환)

    # 각 항목의 y1 좌표 → Excel 행 번호 계산 (빈 행으로 넘버링 정렬)
    assigned_rows = []
    used_rows = {1}  # 헤더 행 예약
    for item in items:
        target = 1 + max(1, round(item['y1'] / PX_PER_ROW))
        while target in used_rows:
            target += 1
        assigned_rows.append(target)
        used_rows.add(target)

    total_data_rows = max(assigned_rows) if assigned_rows else 0

    # 이미지 크기 미리 계산 → A-F열 너비를 이미지 너비에 정확히 맞춤
    header_h_px = int(28 / 0.75)
    data_h_px = int(total_data_rows * ROW_HEIGHT_PT / 0.75)
    ann_h_px = header_h_px + data_h_px
    img_scale = ann_h_px / max(1, annotated_img.height)
    ann_img_w_px = int(annotated_img.width * img_scale)
    # px → Excel 컬럼 너비 단위 변환 (1 char ≈ 7px, 최소 8)
    img_col_width = max(8, ann_img_w_px / 7 / IMAGE_COLS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "번역표"

    header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # A-F열: 이미지 영역 — 너비를 이미지 실제 폭에 맞게 설정
    for c in range(1, IMAGE_COLS + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = img_col_width

    # G열~: 데이터 컬럼 헤더 (#, 위치, 원문, 번역×언어수, 확인)
    NUM_COL  = IMAGE_COLS + 1
    CROP_COL = IMAGE_COLS + 2
    KO_COL   = IMAGE_COLS + 3
    LANG_COL_START = IMAGE_COLS + 4               # 첫 번역 컬럼
    CHK_COL  = LANG_COL_START + len(target_langs) # 번역 컬럼들 뒤 확인 컬럼

    col_defs = [
        (NUM_COL,  '#',             5),
        (CROP_COL, '위치',          30),
        (KO_COL,   '원문 (한국어)', 35),
    ]
    for j, lang in enumerate(target_langs):
        col_defs.append((LANG_COL_START + j, f'번역 ({LANG_NAMES.get(lang, lang)})', 35))
    col_defs.append((CHK_COL, '확인', 8))
    ws.row_dimensions[1].height = 28
    for col_idx, header, width in col_defs:
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # 모든 데이터 행 높이 설정 (빈 행 포함)
    for r in range(2, total_data_rows + 4):
        ws.row_dimensions[r].height = ROW_HEIGHT_PT

    temp_dir = tempfile.mkdtemp()
    try:
        for i, (item, row) in enumerate(zip(items, assigned_rows), 1):
            _, hex6 = COLORS[(i - 1) % len(COLORS)]

            # # 셀
            num_cell = ws.cell(row=row, column=NUM_COL, value=i)
            num_cell.font = Font(bold=True, color='FFFFFF', size=11)
            num_cell.fill = PatternFill(start_color=hex6, end_color=hex6, fill_type='solid')
            num_cell.alignment = Alignment(horizontal='center', vertical='center')
            num_cell.border = border

            # 크롭 이미지
            crop_path = os.path.join(temp_dir, f"crop_{i}.jpg")
            cw, ch = _make_crop(orig_img, item, crop_path)
            xl_crop = XLImage(crop_path)
            xl_crop.width, xl_crop.height = cw, ch
            crop_col = openpyxl.utils.get_column_letter(CROP_COL)
            ws.add_image(xl_crop, f'{crop_col}{row}')
            ws.cell(row=row, column=CROP_COL).border = border

            # 원문
            ko_cell = ws.cell(row=row, column=KO_COL, value=item['text'])
            ko_cell.alignment = Alignment(wrap_text=True, vertical='center')
            ko_cell.border = border

            # 언어별 번역 (선택한 언어 수만큼 컬럼)
            translations = item.get('translations', {})
            for j, lang in enumerate(target_langs):
                tr_cell = ws.cell(row=row, column=LANG_COL_START + j, value=translations.get(lang, ''))
                tr_cell.alignment = Alignment(wrap_text=True, vertical='center')
                tr_cell.border = border

            # 확인
            chk_cell = ws.cell(row=row, column=CHK_COL, value='☐')
            chk_cell.alignment = Alignment(horizontal='center', vertical='center')
            chk_cell.border = border

        # 전체 주석 이미지를 A1에 배치 — 미리 계산한 크기 그대로 사용
        ann_path = os.path.join(temp_dir, "annotated_full.jpg")
        annotated_img.save(ann_path, quality=90)
        xl_full = XLImage(ann_path)
        xl_full.width = ann_img_w_px
        xl_full.height = ann_h_px
        ws.add_image(xl_full, 'A1')

        wb.save(output)
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


# ──────────────────────────────────────────
# 5. 제품 단위 처리
# ──────────────────────────────────────────

def filename_sort_key(name: str):
    """파일명에서 숫자를 순서대로 추출해 정렬키(튜플)로 변환
    예: 03-1.jpg → (3, 1),  01.jpg → (1,),  제품명_2.jpg → (2,)"""
    numbers = re.findall(r'\d+', Path(name).stem)
    return tuple(int(n) for n in numbers) if numbers else (0,)


def get_sorted_images(product_dir: Path) -> list:
    """제품 폴더 내 이미지(GIF 제외)를 파일명 숫자 순서로 정렬"""
    files = [f for f in product_dir.iterdir() if f.suffix.lower() in IMAGE_EXTS]
    return sorted(files, key=lambda p: filename_sort_key(p.name))


def run_pipeline(images: list, product_name: str, target_langs: list, progress=None) -> dict:
    """한 제품의 이미지들을 OCR→번역→주석합성→엑셀(인메모리)로 처리.

    images:       [(파일명, PIL.Image), ...]  (파일명 숫자순으로 내부 정렬)
    target_langs: 번역할 DeepL 언어코드 리스트 (예: ['JA', 'ZH'])
    progress:     선택적 콜백 progress(msg: str) — 진행 상황 표시용

    반환: {'excel': bytes, 'annotated': bytes, 'count': int}
          번역할 텍스트가 없으면 None.
    """
    def log(msg):
        print(msg)
        if progress:
            progress(msg)

    images = sorted(images, key=lambda t: filename_sort_key(t[0]))

    all_items = []
    pil_images = []
    y_offset = 0

    for fname, pil_img in images:
        pil_img = pil_img.convert('RGB')
        pil_images.append(pil_img)

        log(f"  [{fname}] OCR 중...")
        fields = ocr_image(pil_img)
        items = merge_paragraph_lines(filter_korean_only(group_into_lines(fields)))

        for item in items:
            item['y1'] += y_offset
            item['y2'] += y_offset

        log(f"      → {len(items)}개 (한국어 포함 항목만)")
        all_items.extend(items)
        y_offset += pil_img.height

    if not all_items:
        return None

    lang_label = ', '.join(LANG_NAMES.get(l, l) for l in target_langs)
    log(f"  번역 중... (총 {len(all_items)}개 · {lang_label})")
    texts = [item['text'] for item in all_items]
    all_translations = translate_all_langs(texts, target_langs)
    for i, item in enumerate(all_items):
        item['translations'] = {lang: all_translations[lang][i] for lang in all_translations}

    log("  이미지 합성 및 주석 작업 중...")
    combined_img = combine_vertically(pil_images)
    annotated_img = create_annotated_image(combined_img, all_items)

    ann_buf = io.BytesIO()
    annotated_img.save(ann_buf, format='JPEG', quality=90)

    log("  Excel 생성 중...")
    xls_buf = io.BytesIO()
    create_excel(all_items, combined_img, annotated_img, xls_buf, target_langs)

    return {
        'excel': xls_buf.getvalue(),
        'annotated': ann_buf.getvalue(),
        'count': len(all_items),
    }


def process_product(product_dir: Path, output_base: Path):
    """제품 폴더 전체 처리 → 단일 Excel 출력 (CLI 배치)"""
    product_name = product_dir.name
    image_files = get_sorted_images(product_dir)

    if not image_files:
        print(f"  [건너뜀] 이미지 없음: {product_dir}")
        return

    print(f"\n{'='*50}")
    print(f"제품: {product_name}  ({len(image_files)}장)")

    target_langs = getattr(config, 'TARGET_LANGS', ['JA'])
    images = [(p.name, Image.open(p)) for p in image_files]
    result = run_pipeline(images, product_name, target_langs)

    if not result:
        print(f"  [건너뜀] 번역할 텍스트 없음")
        return

    product_out = output_base / product_name
    product_out.mkdir(parents=True, exist_ok=True)
    (product_out / f"{product_name}_annotated.jpg").write_bytes(result['annotated'])
    excel_path = product_out / f"{product_name}_translation.xlsx"
    excel_path.write_bytes(result['excel'])

    print(f"  완료 → {excel_path}")


# ──────────────────────────────────────────
# 메인
# ──────────────────────────────────────────

def main_batch(input_dir: str = 'input', output_dir: str = 'output'):
    """input/ 하위 제품 폴더를 모두 처리"""
    input_path = Path(input_dir)
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)

    all_dirs = sorted([d for d in input_path.iterdir() if d.is_dir()])

    if not all_dirs:
        print("input 폴더 안에 제품 폴더가 없습니다.")
        print("구조 예시: input/제품명/01.jpg")
        return

    # 시작 전에 완료/대기 분류
    done = [d for d in all_dirs if (output_path / d.name / f"{d.name}_translation.xlsx").exists()]
    pending = [d for d in all_dirs if d not in done]

    print(f"전체 {len(all_dirs)}개 제품")
    print(f"  대기: {len(pending)}개  /  완료(건너뜀): {len(done)}개")

    if done:
        print(f"  건너뜀 목록: {', '.join(d.name for d in done)}")

    if not pending:
        print("\n처리할 제품이 없습니다.")
        return

    print(f"\n처리 시작: {', '.join(d.name for d in pending)}\n")

    for product_dir in pending:
        try:
            process_product(product_dir, output_path)
        except Exception as e:
            import traceback
            print(f"  [오류] {product_dir.name}: {e}")
            traceback.print_exc()

    print(f"\n{'='*50}")
    print("전체 완료!")


def main_single(image_path: str):
    """단일 이미지 처리 (레거시 지원)"""
    p = Path(image_path)
    if not p.exists():
        print(f"파일 없음: {image_path}")
        sys.exit(1)

    print(f"[단일 모드] {p.name}")
    target_langs = getattr(config, 'TARGET_LANGS', ['JA'])
    result = run_pipeline([(p.name, Image.open(p))], p.stem, target_langs)

    if not result:
        print("번역할 텍스트 없음")
        return

    output_dir = p.parent.parent.parent / 'output' / p.parent.name
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / f"{p.stem}_annotated.jpg").write_bytes(result['annotated'])
    excel_path = output_dir / f"{p.stem}_translation.xlsx"
    excel_path.write_bytes(result['excel'])

    print(f"\n완료 → {excel_path}")


if __name__ == "__main__":
    script_dir = Path(__file__).parent
    os.chdir(script_dir)
    if len(sys.argv) >= 2:
        main_single(sys.argv[1])
    else:
        main_batch()
